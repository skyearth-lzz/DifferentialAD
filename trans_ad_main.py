import numpy as np
import pandas as pd
from openpyxl.reader.excel import load_workbook
from torch.utils.data import DataLoader, TensorDataset
from tqdm import tqdm

from model.trans_ad.merlin import *
from model.trans_ad.models import *
from model.trans_ad.plotting import *

# from beepy import beep

plt.rcParams['font.sans-serif']=['SimHei']
plt.rcParams['axes.unicode_minus']=False

# 检查CUDA可用性
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

def convert_to_windows(data, model):
    """
    根据w_size使用活动窗口，将各个窗口堆叠起来
    Parameters
    ----------
    data
    model

    Returns
    -------

    """
    windows = []; w_size = model.n_window
    for i, g in enumerate(data):
        if i >= w_size: w = data[i-w_size:i]
        # Repeats this tensor along the specified dimensions. 0 dimension重复w_size-i, 1 dimension重复1（不重复）
        else: w = torch.cat([data[0].repeat(w_size-i, 1), data[0:i]])
        windows.append(w if 'TranAD' in args.model or 'Attention' in args.model else w.view(-1))
    return torch.stack(windows)

def load_dataset(dataset):
    """
    加载数据集，并返回pytorch的Dataloader
    Parameters
    ----------
    dataset: 数据集名称
    数据集名称
    Returns
    -------
    train_loader, test_loader, labels
    """
    folder = os.path.join(output_folder, dataset)
    if not os.path.exists(folder):
        raise Exception('Processed Data not found.')
    loader = []
    for file in ['train', 'test', 'labels']:
        if dataset == 'SMD': file = 'machine-1-1_' + file
        if dataset == 'SMAP': file = 'P-1_' + file
        if dataset == 'MSL': file = 'C-1_' + file
        # 仅使用UCR数据集4个序列中的一个
        if dataset == 'UCR': file = '136_' + file
        if dataset == 'NAB': file = 'ec2_request_latency_system_failure_' + file
        loader.append(np.load(os.path.join(folder, f'{file}.npy')))
    # loader = [i[:, debug:debug+1] for i in loader]
    # 如果命令行参数中包含less，仅使用20%的训练数据
    if args.less: loader[0] = cut_array(0.2, loader[0])

    # 以整个时间序列的时间戳个数为batch_size构建DataLoader
    train_loader = DataLoader(loader[0], batch_size=loader[0].shape[0])
    test_loader = DataLoader(loader[1], batch_size=loader[1].shape[0])
    labels = loader[2]
    return train_loader, test_loader, labels

def save_model(model, optimizer, scheduler, epoch, accuracy_list):
    folder = f'checkpoints/{args.model}_{args.dataset}/'
    os.makedirs(folder, exist_ok=True)
    file_path = f'{folder}/model.ckpt'
    torch.save({
        'epoch': epoch,
        'model_state_dict': model.state_dict(),
        'optimizer_state_dict': optimizer.state_dict(),
        'scheduler_state_dict': scheduler.state_dict(),
        'accuracy_list': accuracy_list}, file_path)

def load_model(modelname, dims):
    import model.trans_ad.models
    # 用于动态获取对象的属性或方法。它的作用是从对象中获取指定名称的属性值，
    # 如果属性不存在，可以返回一个默认值或抛出 AttributeError 异常。
    model_class = getattr(model.trans_ad.models, modelname)
    model = model_class(dims).double().to(device)
    optimizer = torch.optim.AdamW(model.parameters() , lr=model.lr, weight_decay=1e-5)
    # 根据step_size衰减学习率
    scheduler = torch.optim.lr_scheduler.StepLR(optimizer, 5, 0.9)
    fname = f'checkpoints/{args.model}_{args.dataset}/model.ckpt'
    if os.path.exists(fname) and (not args.retrain or args.test):
        # 模型参数文件已存在且非训练环境时
        print(f"{color.GREEN}Loading pre-trained model: {model.name}{color.ENDC}")
        # 加载模型参数
        checkpoint = torch.load(fname)
        model.load_state_dict(checkpoint['model_state_dict'])
        optimizer.load_state_dict(checkpoint['optimizer_state_dict'])
        scheduler.load_state_dict(checkpoint['scheduler_state_dict'])
        epoch = checkpoint['epoch']
        accuracy_list = checkpoint['accuracy_list']
    else:
        print(f"{color.GREEN}Creating new model: {model.name}{color.ENDC}")
        epoch = -1; accuracy_list = []
    return model, optimizer, scheduler, epoch, accuracy_list

def backprop(epoch, model, data, dataO, optimizer, scheduler, training = True):
    """
    模型的具体训练和测试过程
    Parameters
    ----------
    epoch
    model
    data: 经过活动窗口处理的时间序列数据
    dataO：原始的时间序列数据
    optimizer
    scheduler
    training

    Returns
    -------
    Training：
    [average_loss, learning_rate]
    Test:
    [loss, reconstruction_series]

    """
    # 'mean': the sum of the output will be divided by the number of elements in the output
    l = nn.MSELoss(reduction = 'mean' if training else 'none')
    feats = dataO.shape[1]
    if 'DAGMM' in model.name:
        l = nn.MSELoss(reduction = 'none')
        compute = ComputeLoss(model, 0.1, 0.005, 'cpu', model.n_gmm)
        n = epoch + 1; w_size = model.n_window
        l1s = []; l2s = []
        if training:
            for d in data:
                _, x_hat, z, gamma = model(d)
                l1, l2 = l(x_hat, d), l(gamma, d)
                l1s.append(torch.mean(l1).item()); l2s.append(torch.mean(l2).item())
                loss = torch.mean(l1) + torch.mean(l2)
                optimizer.zero_grad()
                loss.backward()
                optimizer.step()
            scheduler.step()
            tqdm.write(f'Epoch {epoch},\tL1 = {np.mean(l1s)},\tL2 = {np.mean(l2s)}')
            return np.mean(l1s)+np.mean(l2s), optimizer.param_groups[0]['lr']
        else:
            ae1s = []
            for d in data:
                _, x_hat, _, _ = model(d)
                ae1s.append(x_hat)
            ae1s = torch.stack(ae1s)
            y_pred = ae1s[:, data.shape[1]-feats:data.shape[1]].view(-1, feats)
            loss = l(ae1s, data)[:, data.shape[1]-feats:data.shape[1]].view(-1, feats)
            return loss.detach().numpy(), y_pred.detach().numpy()
    if 'Attention' in model.name:
        l = nn.MSELoss(reduction = 'none')
        n = epoch + 1; w_size = model.n_window
        l1s = []; res = []
        if training:
            for d in data:
                ae, ats = model(d)
                # res.append(torch.mean(ats, axis=0).view(-1))
                l1 = l(ae, d)
                l1s.append(torch.mean(l1).item())
                loss = torch.mean(l1)
                optimizer.zero_grad()
                loss.backward()
                optimizer.step()
            # res = torch.stack(res); np.save('ascores.npy', res.detach().numpy())
            scheduler.step()
            tqdm.write(f'Epoch {epoch},\tL1 = {np.mean(l1s)}')
            return np.mean(l1s), optimizer.param_groups[0]['lr']
        else:
            ae1s, y_pred = [], []
            for d in data:
                ae1 = model(d)
                y_pred.append(ae1[-1])
                ae1s.append(ae1)
            ae1s, y_pred = torch.stack(ae1s), torch.stack(y_pred)
            loss = torch.mean(l(ae1s, data), axis=1)
            return loss.detach().numpy(), y_pred.detach().numpy()
    elif 'OmniAnomaly' in model.name:
        if training:
            mses, klds = [], []
            for i, d in enumerate(data):
                y_pred, mu, logvar, hidden = model(d, hidden if i else None)
                MSE = l(y_pred, d)
                KLD = -0.5 * torch.sum(1 + logvar - mu.pow(2) - logvar.exp(), dim=0)
                loss = MSE + model.beta * KLD
                mses.append(torch.mean(MSE).item()); klds.append(model.beta * torch.mean(KLD).item())
                optimizer.zero_grad()
                loss.backward()
                optimizer.step()
            tqdm.write(f'Epoch {epoch},\tMSE = {np.mean(mses)},\tKLD = {np.mean(klds)}')
            scheduler.step()
            return loss.item(), optimizer.param_groups[0]['lr']
        else:
            y_preds = []
            for i, d in enumerate(data):
                y_pred, _, _, hidden = model(d, hidden if i else None)
                y_preds.append(y_pred)
            y_pred = torch.stack(y_preds)
            MSE = l(y_pred, data)
            return MSE.detach().numpy(), y_pred.detach().numpy()
    elif 'USAD' in model.name:
        l = nn.MSELoss(reduction = 'none')
        n = epoch + 1; w_size = model.n_window
        l1s, l2s = [], []
        if training:
            for d in data:
                ae1s, ae2s, ae2ae1s = model(d)
                l1 = (1 / n) * l(ae1s, d) + (1 - 1/n) * l(ae2ae1s, d)
                l2 = (1 / n) * l(ae2s, d) - (1 - 1/n) * l(ae2ae1s, d)
                l1s.append(torch.mean(l1).item()); l2s.append(torch.mean(l2).item())
                loss = torch.mean(l1 + l2)
                optimizer.zero_grad()
                loss.backward()
                optimizer.step()
            scheduler.step()
            tqdm.write(f'Epoch {epoch},\tL1 = {np.mean(l1s)},\tL2 = {np.mean(l2s)}')
            return np.mean(l1s)+np.mean(l2s), optimizer.param_groups[0]['lr']
        else:
            ae1s, ae2s, ae2ae1s = [], [], []
            for d in data:
                ae1, ae2, ae2ae1 = model(d)
                ae1s.append(ae1); ae2s.append(ae2); ae2ae1s.append(ae2ae1)
            ae1s, ae2s, ae2ae1s = torch.stack(ae1s), torch.stack(ae2s), torch.stack(ae2ae1s)
            y_pred = ae1s[:, data.shape[1]-feats:data.shape[1]].view(-1, feats)
            loss = 0.1 * l(ae1s, data) + 0.9 * l(ae2ae1s, data)
            loss = loss[:, data.shape[1]-feats:data.shape[1]].view(-1, feats)
            return loss.detach().numpy(), y_pred.detach().numpy()
    elif model.name in ['GDN', 'MTAD_GAT', 'MSCRED', 'CAE_M']:
        l = nn.MSELoss(reduction = 'none')
        n = epoch + 1; w_size = model.n_window
        l1s = []
        if training:
            for i, d in enumerate(data):
                if 'MTAD_GAT' in model.name:
                    x, h = model(d, h if i else None)
                else:
                    x = model(d)
                loss = torch.mean(l(x, d))
                l1s.append(torch.mean(loss).item())
                optimizer.zero_grad()
                loss.backward()
                optimizer.step()
            tqdm.write(f'Epoch {epoch},\tMSE = {np.mean(l1s)}')
            return np.mean(l1s), optimizer.param_groups[0]['lr']
        else:
            xs = []
            for d in data:
                if 'MTAD_GAT' in model.name:
                    x, h = model(d, None)
                else:
                    x = model(d)
                xs.append(x)
            xs = torch.stack(xs)
            y_pred = xs[:, data.shape[1]-feats:data.shape[1]].view(-1, feats)
            loss = l(xs, data)
            loss = loss[:, data.shape[1]-feats:data.shape[1]].view(-1, feats)
            return loss.detach().numpy(), y_pred.detach().numpy()
    elif 'GAN' in model.name:
        l = nn.MSELoss(reduction = 'none')
        bcel = nn.BCELoss(reduction = 'mean')
        msel = nn.MSELoss(reduction = 'mean')
        real_label, fake_label = torch.tensor([0.9]), torch.tensor([0.1]) # label smoothing
        real_label, fake_label = real_label.type(torch.DoubleTensor), fake_label.type(torch.DoubleTensor)
        n = epoch + 1; w_size = model.n_window
        mses, gls, dls = [], [], []
        if training:
            for d in data:
                # training discriminator
                model.discriminator.zero_grad()
                _, real, fake = model(d)
                dl = bcel(real, real_label) + bcel(fake, fake_label)
                dl.backward()
                model.generator.zero_grad()
                optimizer.step()
                # training generator
                z, _, fake = model(d)
                mse = msel(z, d)
                gl = bcel(fake, real_label)
                tl = gl + mse
                tl.backward()
                model.discriminator.zero_grad()
                optimizer.step()
                mses.append(mse.item()); gls.append(gl.item()); dls.append(dl.item())
                # tqdm.write(f'Epoch {epoch},\tMSE = {mse},\tG = {gl},\tD = {dl}')
            tqdm.write(f'Epoch {epoch},\tMSE = {np.mean(mses)},\tG = {np.mean(gls)},\tD = {np.mean(dls)}')
            return np.mean(gls)+np.mean(dls), optimizer.param_groups[0]['lr']
        else:
            outputs = []
            for d in data:
                z, _, _ = model(d)
                outputs.append(z)
            outputs = torch.stack(outputs)
            y_pred = outputs[:, data.shape[1]-feats:data.shape[1]].view(-1, feats)
            loss = l(outputs, data)
            loss = loss[:, data.shape[1]-feats:data.shape[1]].view(-1, feats)
            return loss.detach().numpy(), y_pred.detach().numpy()
    elif 'TranAD' in model.name:
        # reduction='none'表示不对损失进行任何缩减操作,返回每个元素的损失值
        # 这样可以得到更细粒度的损失信息,便于后续分析每个样本的重建误差
        l = nn.MSELoss(reduction = 'none')
        data_x = torch.DoubleTensor(data); dataset = TensorDataset(data_x, data_x)
        # 设置训练的batch size
        bs = model.batch if training else len(data)
        dataloader = DataLoader(dataset, batch_size = bs)
        n = epoch + 1; w_size = model.n_window
        l1s, l2s = [], []
        model.train()
        if training:
            for d, _ in dataloader:
                d = d.to(device)
                local_bs = d.shape[0]
                # 维度变为w_size * batch_size * feats
                window = d.permute(1, 0, 2)
                # elem就是window i在原有时间序列时间为i的值
                elem = window[-1, :, :].view(1, local_bs, feats)
                z = model(window, elem)
                l1 = l(z, elem) if not isinstance(z, tuple) else (1 / n) * l(z[0], elem) + (1 - 1/n) * l(z[1], elem)
                if isinstance(z, tuple): z = z[1]
                l1s.append(torch.mean(l1).item())
                loss = torch.mean(l1)
                optimizer.zero_grad()
                loss.backward(retain_graph=True)
                optimizer.step()
            # 调整学习率
            scheduler.step()
            tqdm.write(f'Epoch {epoch},\tL1 = {np.mean(l1s)}')
            return np.mean(l1s), optimizer.param_groups[0]['lr']
        else:
            for d, _ in dataloader:
                d = d.to(device)
                window = d.permute(1, 0, 2)
                elem = window[-1, :, :].view(1, bs, feats)
                z = model(window, elem)
                if isinstance(z, tuple): z = z[1]
            loss = l(z, elem)[0]
            return loss.detach().cpu().numpy(), z.detach().cpu().numpy()[0]
    else:
        y_pred = model(data)
        loss = l(y_pred, data)
        if training:
            tqdm.write(f'Epoch {epoch},\tMSE = {loss}')
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
            scheduler.step()
            return loss.item(), optimizer.param_groups[0]['lr']
        else:
            return loss.detach().numpy(), y_pred.detach().numpy()


if __name__ == '__main__1':
    train_loader, test_loader, labels = load_dataset(args.dataset)
    if args.model in ['MERLIN']:
        eval(f'run_{args.model.lower()}(test_loader, labels, args.dataset)')
    model, optimizer, scheduler, epoch, accuracy_list = load_model(args.model, labels.shape[1])

    ## Prepare data
    trainD, testD = next(iter(train_loader)), next(iter(test_loader))
    trainO, testO = trainD, testD
    if model.name in ['Attention', 'DAGMM', 'USAD', 'MSCRED', 'CAE_M', 'GDN', 'MTAD_GAT',
                      'MAD_GAN'] or 'TranAD' in model.name:
        trainD, testD = convert_to_windows(trainD, model), convert_to_windows(testD, model)

    print(f'{color.HEADER}Training {args.model} on {args.dataset}{color.ENDC}')
    num_epochs = args.epoch

    metrics = []

    for epoch in range(num_epochs):
        model.train()
        lossT, lr = backprop(epoch, model, trainD, trainO, optimizer, scheduler)
        save_model(model, optimizer, scheduler, epoch, accuracy_list)

        ### Testing phase
        torch.zero_grad = True
        model.eval()
        loss, y_pred = backprop(0, model, testD, testO, optimizer, scheduler, training=False)


        # 计算训练集数据在训练好的模型上的损失
        lossT, _ = backprop(0, model, trainD, trainO, optimizer, scheduler, training=False)
        for i in range(loss.shape[1]):
            # 取出每个维度的训练损失，测试损失，异常标签
            lt, l, ls = lossT[:, i], loss[:, i], labels[:, i]
            result, pred = pot_eval(lt, l, ls);
            preds.append(pred)
        # preds = np.concatenate([i.reshape(-1, 1) + 0 for i in preds], axis=1)
        # pd.DataFrame(preds, columns=[str(i) for i in range(10)]).to_csv('labels.csv')
        lossTfinal, lossFinal = np.mean(lossT, axis=1), np.mean(loss, axis=1)
        labelsFinal = (np.sum(labels, axis=1) >= 1) + 0
        result, _ = pot_eval(lossTfinal, lossFinal, labelsFinal)
        result.update(hit_att(loss, labels))
        result.update(ndcg(loss, labels))
        result['train_loss'] = np.mean(lossTfinal)
        result['test_loss'] = np.mean(lossFinal)
        metrics.append(result)
        # pprint(getresults2(df, result))
        # beep(4)


    data_frame = pd.DataFrame(metrics)
    # 检查文件是否已存在
    try:
        # 加载现有的 Excel 文件
        book = load_workbook("metrics.xlsx")
        if args.dataset in book.sheetnames:
            del book[args.dataset]

        # 创建一个新的工作表并写入数据
        with pd.ExcelWriter("metrics.xlsx", engine="openpyxl", mode="a") as writer:
            data_frame.to_excel(writer, sheet_name=args.dataset, index=True)
    except FileNotFoundError:
        # 如果文件不存在，创建一个新的 ExcelWriter
        with pd.ExcelWriter("metrics.xlsx", engine="openpyxl", mode="w") as writer:
            data_frame.to_excel(writer, sheet_name=args.dataset, index=True)


if __name__ == '__main__':
    train_loader, test_loader, labels = load_dataset(args.dataset)
    if args.model in ['MERLIN']:
        eval(f'run_{args.model.lower()}(test_loader, labels, args.dataset)')
    model, optimizer, scheduler, epoch, accuracy_list = load_model(args.model, labels.shape[1])

    ## Prepare data
    trainD, testD = next(iter(train_loader)), next(iter(test_loader))
    trainO, testO = trainD, testD
    if model.name in ['Attention', 'DAGMM', 'USAD', 'MSCRED', 'CAE_M', 'GDN', 'MTAD_GAT', 'MAD_GAN'] or 'TranAD' in model.name:
        trainD, testD = convert_to_windows(trainD, model), convert_to_windows(testD, model)

    ### Training phase
    if not args.test:
        print(f'{color.HEADER}Training {args.model} on {args.dataset}{color.ENDC}')
        num_epochs = args.epoch; e = epoch + 1; start = time()
        for e in tqdm(list(range(epoch+1, epoch+num_epochs+1))):
            lossT, lr = backprop(e, model, trainD, trainO, optimizer, scheduler)
            accuracy_list.append((lossT, lr))
        print(color.BOLD+'Training time: '+"{:10.4f}".format(time()-start)+' s'+color.ENDC)
        save_model(model, optimizer, scheduler, e, accuracy_list)
        plot_accuracies(accuracy_list, f'{args.model}_{args.dataset}')

    ### Testing phase
    torch.zero_grad = True
    model.eval()
    print(f'{color.HEADER}Testing {args.model} on {args.dataset}{color.ENDC}')
    loss, y_pred = backprop(0, model, testD, testO, optimizer, scheduler, training=False)

    ### Plot curves
    if not args.test:
        # 原始时间序列沿着0维度（时间维度）滚动1位
        if 'TranAD' in model.name: testO = torch.roll(testO, 1, 0)
        plotter(f'{args.model}_{args.dataset}', testO, y_pred, loss, labels)

    ### Scores
    df = pd.DataFrame()

    # 计算训练集数据在训练好的模型上的损失
    lossT, _ = backprop(0, model, trainD, trainO, optimizer, scheduler, training=False)
    for i in range(loss.shape[1]):
        # 取出每个维度的训练损失，测试损失，异常标签
        lt, l, ls = lossT[:, i], loss[:, i], labels[:, i]
        result, pred = pot_eval(lt, l, ls); preds.append(pred)
        df = df._append(result, ignore_index=True)
    # preds = np.concatenate([i.reshape(-1, 1) + 0 for i in preds], axis=1)
    # pd.DataFrame(preds, columns=[str(i) for i in range(10)]).to_csv('labels.csv')
    lossTfinal, lossFinal = np.mean(lossT, axis=1), np.mean(loss, axis=1)
    labelsFinal = (np.sum(labels, axis=1) >= 1) + 0
    result, _ = pot_eval(lossTfinal, lossFinal, labelsFinal)
    result.update(hit_att(loss, labels))
    result.update(ndcg(loss, labels))
    print(df)
    pprint(result)
    # pprint(getresults2(df, result))
    # beep(4)
